"""Excel (.xlsx) invoice parser for extracting products and supplier."""
import re
from decimal import Decimal
from pathlib import Path

import openpyxl

from bot.services.pdf_parser import normalize_unit, _parse_number


def _extract_supplier_from_header(ws) -> str | None:
    """Ищет поставщика в первых строках Excel (поставщик / продавец)."""
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        cell_val = str(ws.cell(row_idx, 1).value or "")
        if not cell_val.strip():
            continue
        low = cell_val.lower()
        for marker in ("поставщик", "продавец", "seller"):
            if marker in low:
                match = re.search(
                    r'[:\s]+[«""]?([A-Za-zА-Яа-яЁёA-Za-z0-9\s\-\.]+)',
                    cell_val[low.index(marker) + len(marker):],
                )
                if match:
                    name = match.group(1).strip().strip('"»"«')
                    name = re.sub(r"\s*[;,]?\s*покупатель\s*:.*$", "", name, flags=re.I)
                    if len(name) >= 3:
                        return name[:120]
    return None


def _find_header_row(ws) -> tuple[int, dict] | None:
    """
    Находит строку-заголовок с колонками (наименование, кол-во, цена и т.д.).
    Возвращает (номер строки, словарь индексов колонок) или None.
    """
    for row_idx in range(1, min(ws.max_row + 1, 30)):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            row_vals.append(str(ws.cell(row_idx, c).value or "").lower().strip())

        row_text = " ".join(row_vals)
        has_name = "наименование" in row_text
        has_qty = "кол" in row_text or "количество" in row_text
        non_empty = sum(1 for v in row_vals if v)

        if non_empty >= 3 and has_name and has_qty:
            indices = {
                "num": 0,
                "name": 1,
                "code": 2,
                "unit": 3,
                "quantity": 4,
                "price": 5,
                "cost_vat": 8,
            }
            for i, cell in enumerate(row_vals):
                if cell == "№" or (len(cell) <= 3 and "№" in cell):
                    indices["num"] = i
                elif "наименование" in cell:
                    indices["name"] = i
                elif "ед" in cell or "измер" in cell:
                    indices["unit"] = i
                elif "кол" in cell:
                    indices["quantity"] = i
                elif "цена" in cell and "ндс" not in cell:
                    indices["price"] = i
                elif "стоимость" in cell and ("учетом" in cell or "учётом" in cell) and "ндс" in cell:
                    indices["cost_vat"] = i
                elif "идентификацион" in cell or ("код" in cell and "штрих" not in cell):
                    indices["code"] = i
            return row_idx, indices
    return None


def parse_excel(file_path: str | Path) -> dict:
    """
    Извлекает товары и поставщика из Excel (.xlsx) счёт-фактуры.
    Возвращает: {"products": [...], "supplier_name": str | None}
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    supplier_name = _extract_supplier_from_header(ws)

    result = _find_header_row(ws)
    if result is None:
        return {"products": [], "supplier_name": supplier_name}
    header_row_idx, col = result

    skip_words = {"итого", "всего", "total", "сумма"}
    products: list[dict] = []

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        def cell(c_idx: int) -> str:
            v = ws.cell(row_idx, c_idx + 1).value
            return str(v).strip() if v is not None else ""

        num_val = cell(col["num"])
        if not num_val:
            continue
        try:
            num = int(float(num_val))
            if num <= 0:
                continue
        except (ValueError, TypeError):
            continue

        name = cell(col["name"]).replace("\n", " ")
        name = re.sub(r"\s*\*[\d\w]+\s*$", "", name).strip()
        if not name or any(s in name.lower() for s in skip_words):
            continue
        if re.fullmatch(r"\d+", name):
            continue

        code = cell(col["code"]) if col["code"] < ws.max_column else ""
        unit = normalize_unit(cell(col["unit"]) if col["unit"] < ws.max_column else "")

        qty = _parse_number(cell(col["quantity"]))
        cost_vat = _parse_number(cell(col["cost_vat"]))
        price = _parse_number(cell(col["price"]))

        if qty is None or qty <= 0:
            continue

        if cost_vat and cost_vat > 0:
            price_with_vat = float(cost_vat / qty)
        elif price:
            price_with_vat = float(price)
        else:
            price_with_vat = 0.0

        products.append({
            "name": name,
            "unit": unit,
            "quantity": float(qty),
            "price_with_vat": round(price_with_vat, 2),
            "code_from_pdf": (code or "")[:80],
        })

    return {"products": products, "supplier_name": supplier_name}
